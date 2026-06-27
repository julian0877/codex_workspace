from openpyxl import load_workbook

path = r'e:\codex_workspace\cal\附件1-钢梁板式吊耳验算-【5t^F50t】-双吊耳-201906-.xlsx'
out_path = r'e:\codex_workspace\cal\inspect_excel_output.txt'
wb = load_workbook(path, data_only=True)
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('sheets = ' + repr(wb.sheetnames) + '\n')
    for ws in wb.worksheets:
        f.write('sheet ' + ws.title + '\n')
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
            if any(cell is not None for cell in row):
                f.write(str(i) + ' ' + repr(row) + '\n')
        f.write('--- end sheet ---\n')
